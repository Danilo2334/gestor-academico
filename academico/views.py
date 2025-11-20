from datetime import datetime

import openpyxl
from django import forms
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from .decorators import rol_requerido
from .forms import (
    AsistenciaForm,
    CalificacionForm,
    CursoForm,
    DocenteForm,
    EstudianteForm,
    MateriaForm,
)
from .models import (
    Asistencia,
    Calificacion,
    Curso,
    Docente,
    Estudiante,
    Materia,
    Matricula,
)


def inicio(request):
    if request.user.is_authenticated:
        logout(request)
        messages.info(request, "Tu sesión anterior se cerró. Inicia nuevamente.")
    return redirect("login")


def logout_view(request):
    logout(request)
    messages.info(request, "Has cerrado sesión correctamente.")
    return redirect("login")


@login_required
def dashboard(request):
    """
    Dashboard visible para cualquier usuario autenticado.
    ADMIN verá la foto global, DOCENTE y ESTUDIANTE verán los mismos indicadores,
    pero igual los permisos se controlan en cada vista.
    """
    total_estudiantes = Estudiante.objects.count()
    total_docentes = Docente.objects.count()
    total_cursos = Curso.objects.count()
    total_materias = Materia.objects.count()

    promedio_por_curso = (
        Calificacion.objects.values("materia__curso__nombre")
        .annotate(promedio=Avg("nota"))
        .order_by("materia__curso__nombre")
    )

    hoy = timezone.now().date()
    mes_actual = hoy.month
    asistencia_mes = (
        Asistencia.objects.filter(fecha__month=mes_actual)
        .values("estado")
        .annotate(total=Count("id"))
    )

    context = {
        "total_estudiantes": total_estudiantes,
        "total_docentes": total_docentes,
        "total_cursos": total_cursos,
        "total_materias": total_materias,
        "promedio_por_curso": promedio_por_curso,
        "asistencia_mes": asistencia_mes,
    }
    return render(request, "dashboard.html", context)


@rol_requerido("ADMIN")
def estudiante_list(request):
    q = request.GET.get("q", "")
    estudiantes = Estudiante.objects.all()
    if q:
        estudiantes = (
            estudiantes.filter(user__first_name__icontains=q)
            | estudiantes.filter(user__last_name__icontains=q)
            | estudiantes.filter(documento__icontains=q)
        )
    return render(request, "estudiantes/estudiante_list.html", {"estudiantes": estudiantes, "q": q})


@rol_requerido("ADMIN")
def estudiante_create(request):
    if request.method == "POST":
        form = EstudianteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Estudiante creado correctamente.")
            return redirect("estudiante_list")
    else:
        form = EstudianteForm()
    return render(request, "estudiantes/estudiante_form.html", {"form": form})


@rol_requerido("ADMIN")
def estudiante_update(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    if request.method == "POST":
        form = EstudianteForm(request.POST, instance=estudiante, user_instance=estudiante.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Estudiante actualizado correctamente.")
            return redirect("estudiante_list")
    else:
        form = EstudianteForm(instance=estudiante, user_instance=estudiante.user)
    return render(request, "estudiantes/estudiante_form.html", {"form": form})


@rol_requerido("ADMIN")
def estudiante_delete(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    if request.method == "POST":
        estudiante.delete()
        messages.success(request, "Estudiante eliminado.")
        return redirect("estudiante_list")
    return render(request, "estudiantes/estudiante_confirm_delete.html", {"estudiante": estudiante})


@rol_requerido("ADMIN")
def docente_list(request):
    q = request.GET.get("q", "")
    docentes = Docente.objects.all()
    if q:
        docentes = docentes.filter(user__first_name__icontains=q) | docentes.filter(user__last_name__icontains=q)
    return render(request, "docentes/docente_list.html", {"docentes": docentes, "q": q})


@rol_requerido("ADMIN")
def docente_create(request):
    if request.method == "POST":
        form = DocenteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Docente creado correctamente.")
            return redirect("docente_list")
    else:
        form = DocenteForm()
    return render(request, "docentes/docente_form.html", {"form": form})


@rol_requerido("ADMIN")
def docente_update(request, pk):
    docente = get_object_or_404(Docente, pk=pk)
    if request.method == "POST":
        form = DocenteForm(request.POST, instance=docente, user_instance=docente.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Docente actualizado correctamente.")
            return redirect("docente_list")
    else:
        form = DocenteForm(instance=docente, user_instance=docente.user)
    return render(request, "docentes/docente_form.html", {"form": form})


@rol_requerido("ADMIN")
def docente_delete(request, pk):
    docente = get_object_or_404(Docente, pk=pk)
    if request.method == "POST":
        docente.delete()
        messages.success(request, "Docente eliminado.")
        return redirect("docente_list")
    return render(request, "docentes/docente_confirm_delete.html", {"docente": docente})


@rol_requerido("ADMIN")
def curso_list(request):
    q = request.GET.get("q", "")
    cursos = Curso.objects.all()
    if q:
        cursos = cursos.filter(nombre__icontains=q) | cursos.filter(codigo__icontains=q)
    return render(request, "cursos/curso_list.html", {"cursos": cursos, "q": q})


@rol_requerido("ADMIN")
def curso_create(request):
    if request.method == "POST":
        form = CursoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Curso creado correctamente.")
            return redirect("curso_list")
    else:
        form = CursoForm()
    return render(request, "cursos/curso_form.html", {"form": form})


@rol_requerido("ADMIN")
def curso_update(request, pk):
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == "POST":
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            messages.success(request, "Curso actualizado correctamente.")
            return redirect("curso_list")
    else:
        form = CursoForm(instance=curso)
    return render(request, "cursos/curso_form.html", {"form": form})


@rol_requerido("ADMIN")
def curso_delete(request, pk):
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == "POST":
        curso.delete()
        messages.success(request, "Curso eliminado.")
        return redirect("curso_list")
    return render(request, "cursos/curso_confirm_delete.html", {"curso": curso})


@rol_requerido("ADMIN")
def materia_list(request):
    materias = Materia.objects.select_related("curso", "docente").all()
    return render(request, "materias/materia_list.html", {"materias": materias})


@rol_requerido("ADMIN")
def materia_create(request):
    if request.method == "POST":
        form = MateriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Materia creada correctamente.")
            return redirect("materia_list")
    else:
        form = MateriaForm()
    return render(request, "materias/materia_form.html", {"form": form})


@rol_requerido("ADMIN")
def materia_update(request, pk):
    materia = get_object_or_404(Materia, pk=pk)
    if request.method == "POST":
        form = MateriaForm(request.POST, instance=materia)
        if form.is_valid():
            form.save()
            messages.success(request, "Materia actualizada correctamente.")
            return redirect("materia_list")
    else:
        form = MateriaForm(instance=materia)
    return render(request, "materias/materia_form.html", {"form": form})


@rol_requerido("ADMIN")
def materia_delete(request, pk):
    materia = get_object_or_404(Materia, pk=pk)
    if request.method == "POST":
        materia.delete()
        messages.success(request, "Materia eliminada.")
        return redirect("materia_list")
    return render(request, "materias/materia_confirm_delete.html", {"materia": materia})


@rol_requerido("ADMIN")
def calificacion_list(request):
    calificaciones = Calificacion.objects.select_related(
        "estudiante__user", "materia__curso", "docente__user"
    ).all()
    return render(request, "calificaciones/calificacion_list.html", {"calificaciones": calificaciones})


@rol_requerido("ADMIN")
def calificacion_create(request):
    if request.method == "POST":
        form = CalificacionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Calificación registrada correctamente.")
            return redirect("calificacion_list")
    else:
        form = CalificacionForm()
    return render(request, "calificaciones/calificacion_form.html", {"form": form})


@rol_requerido("ADMIN")
def calificacion_update(request, pk):
    calificacion = get_object_or_404(Calificacion, pk=pk)
    if request.method == "POST":
        form = CalificacionForm(request.POST, instance=calificacion)
        if form.is_valid():
            form.save()
            messages.success(request, "Calificación actualizada correctamente.")
            return redirect("calificacion_list")
    else:
        form = CalificacionForm(instance=calificacion)
    return render(request, "calificaciones/calificacion_form.html", {"form": form})


@rol_requerido("ADMIN")
def calificacion_delete(request, pk):
    calificacion = get_object_or_404(Calificacion, pk=pk)
    if request.method == "POST":
        calificacion.delete()
        messages.success(request, "Calificación eliminada.")
        return redirect("calificacion_list")
    return render(request, "calificaciones/calificacion_confirm_delete.html", {"calificacion": calificacion})


@rol_requerido("ADMIN", "DOCENTE")
def asistencia_list(request):
    asistencias = Asistencia.objects.select_related("estudiante__user", "materia__curso")
    perfil = request.user.perfil
    if perfil.rol == "DOCENTE":
        docente = get_object_or_404(Docente, user=request.user)
        asistencias = asistencias.filter(materia__docente=docente)
    else:
        asistencias = asistencias.all()
    return render(request, "asistencias/asistencia_list.html", {"asistencias": asistencias})


@rol_requerido("ADMIN")
def asistencia_create(request):
    if request.method == "POST":
        form = AsistenciaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Asistencia registrada.")
            return redirect("asistencia_list")
    else:
        form = AsistenciaForm()
    return render(request, "asistencias/asistencia_form.html", {"form": form})


@rol_requerido("ADMIN")
def asistencia_delete(request, pk):
    asistencia = get_object_or_404(Asistencia, pk=pk)
    if request.method == "POST":
        asistencia.delete()
        messages.success(request, "Registro de asistencia eliminado.")
        return redirect("asistencia_list")
    return render(request, "asistencias/asistencia_confirm_delete.html", {"asistencia": asistencia})


@rol_requerido("ADMIN", "DOCENTE")
def marcar_asistencia(request, materia_id):
    materia = get_object_or_404(Materia, pk=materia_id)

    perfil = request.user.perfil
    if perfil.rol == "DOCENTE" and (materia.docente is None or materia.docente.user != request.user):
        messages.error(request, "No puedes registrar asistencia para una materia que no dictas.")
        return redirect("dashboard")

    estudiantes = Estudiante.objects.filter(
        matriculas__curso=materia.curso,
        matriculas__activo=True,
    ).distinct()

    fecha_str = request.POST.get("fecha") if request.method == "POST" else request.GET.get("fecha")
    fecha = None
    if fecha_str:
        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            fecha = None

    if request.method == "POST" and fecha:
        for est in estudiantes:
            estado = request.POST.get(f"estado_{est.id}")
            if estado:
                Asistencia.objects.update_or_create(
                    estudiante=est,
                    materia=materia,
                    fecha=fecha,
                    defaults={"estado": estado},
                )
        messages.success(request, "Asistencia guardada correctamente.")
        return redirect("asistencia_list")

    context = {
        "materia": materia,
        "estudiantes": estudiantes,
        "fecha": fecha_str or "",
    }
    return render(request, "asistencias/marcar_asistencia.html", context)


@rol_requerido("DOCENTE", "ADMIN")
def docente_mis_materias(request):
    """
    Docente ve solo sus materias.
    ADMIN vería todas (útil para pruebas).
    """
    perfil = request.user.perfil

    if perfil.rol == "ADMIN":
        materias = Materia.objects.select_related("curso", "docente").all()
    else:
        docente = get_object_or_404(Docente, user=request.user)
        materias = Materia.objects.filter(docente=docente).select_related("curso")

    return render(request, "docentes/mis_materias.html", {"materias": materias})


@rol_requerido("DOCENTE", "ADMIN")
def docente_calificaciones_materia(request, materia_id):
    """
    Docente registra y consulta calificaciones solo de SU materia.
    ADMIN puede usarlo para pruebas.
    """
    materia = get_object_or_404(Materia, pk=materia_id)

    perfil = request.user.perfil
    if perfil.rol == "DOCENTE":
        if materia.docente is None or materia.docente.user != request.user:
            messages.error(request, "No puedes gestionar calificaciones de una materia que no dictas.")
            return redirect("dashboard")

    calificaciones = Calificacion.objects.filter(materia=materia).select_related("estudiante__user")

    if request.method == "POST":
        form = CalificacionForm(request.POST)
        form.fields["estudiante"].queryset = Estudiante.objects.filter(
            matriculas__curso=materia.curso,
            matriculas__activo=True,
        ).distinct()
        if form.is_valid():
            cal = form.save(commit=False)
            cal.materia = materia
            if perfil.rol == "DOCENTE":
                cal.docente = get_object_or_404(Docente, user=request.user)
            cal.save()
            messages.success(request, "Calificación registrada correctamente.")
            return redirect("docente_calificaciones_materia", materia_id=materia.id)
    else:
        form = CalificacionForm()
        form.fields["materia"].widget = forms.HiddenInput()
        form.fields["docente"].widget = forms.HiddenInput()
        form.fields["estudiante"].queryset = Estudiante.objects.filter(
            matriculas__curso=materia.curso,
            matriculas__activo=True,
        ).distinct()

    context = {
        "materia": materia,
        "calificaciones": calificaciones,
        "form": form,
    }
    return render(request, "docentes/calificaciones_materia.html", context)


@rol_requerido("ESTUDIANTE")
def mis_calificaciones(request):
    estudiante = get_object_or_404(Estudiante, user=request.user)
    calificaciones = Calificacion.objects.filter(estudiante=estudiante).select_related(
        "materia__curso", "docente__user"
    )
    return render(request, "calificaciones/mis_calificaciones.html", {"calificaciones": calificaciones})


@rol_requerido("ESTUDIANTE")
def mi_asistencia(request):
    estudiante = get_object_or_404(Estudiante, user=request.user)
    asistencias = Asistencia.objects.filter(estudiante=estudiante).select_related("materia__curso")
    return render(request, "asistencias/mi_asistencia.html", {"asistencias": asistencias})


@rol_requerido("ESTUDIANTE")
def mi_boletin(request):
    estudiante = get_object_or_404(Estudiante, user=request.user)
    return boletin_pdf(request, estudiante.id)


@rol_requerido("ADMIN", "ESTUDIANTE")
def boletin_pdf(request, estudiante_id):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)

    perfil = request.user.perfil
    if perfil.rol == "ESTUDIANTE" and estudiante.user != request.user:
        messages.error(request, "No puedes ver el boletín de otro estudiante.")
        return redirect("dashboard")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="boletin_{estudiante.documento}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 50, "Boletín de calificaciones")

    p.setFont("Helvetica", 12)
    p.drawString(50, height - 80, f"Estudiante: {estudiante.user.get_full_name()}")
    p.drawString(50, height - 100, f"Documento: {estudiante.documento}")

    y = height - 140
    p.drawString(50, y, "Materia")
    p.drawString(300, y, "Periodo")
    p.drawString(400, y, "Nota")

    y -= 20
    calificaciones = estudiante.calificaciones.select_related("materia").all()
    for cal in calificaciones:
        if y < 50:
            p.showPage()
            y = height - 50
        p.drawString(50, y, cal.materia.nombre)
        p.drawString(300, y, cal.periodo)
        p.drawString(400, y, str(cal.nota))
        y -= 20

    p.showPage()
    p.save()
    return response


@rol_requerido("ADMIN")
def calificaciones_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calificaciones"

    headers = ["Estudiante", "Documento", "Curso", "Materia", "Periodo", "Nota"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    calificaciones = Calificacion.objects.select_related("estudiante__user", "materia__curso").all()
    row_num = 2
    for cal in calificaciones:
        ws[f"A{row_num}"] = cal.estudiante.user.get_full_name()
        ws[f"B{row_num}"] = cal.estudiante.documento
        ws[f"C{row_num}"] = cal.materia.curso.nombre
        ws[f"D{row_num}"] = cal.materia.nombre
        ws[f"E{row_num}"] = cal.periodo
        ws[f"F{row_num}"] = float(cal.nota)
        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="calificaciones.xlsx"'
    wb.save(response)
    return response


@rol_requerido("ADMIN")
def asistencia_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    headers = ["Estudiante", "Documento", "Curso", "Materia", "Fecha", "Estado"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    asistencias = Asistencia.objects.select_related("estudiante__user", "materia__curso").all()
    row_num = 2
    for a in asistencias:
        ws[f"A{row_num}"] = a.estudiante.user.get_full_name()
        ws[f"B{row_num}"] = a.estudiante.documento
        ws[f"C{row_num}"] = a.materia.curso.nombre
        ws[f"D{row_num}"] = a.materia.nombre
        ws[f"E{row_num}"] = a.fecha.strftime("%Y-%m-%d")
        ws[f"F{row_num}"] = a.get_estado_display()
        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="asistencia.xlsx"'
    wb.save(response)
    return response
